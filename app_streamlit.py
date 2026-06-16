import os
import re
import io
import pandas as pd
import requests
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Set premium page config
st.set_page_config(
    page_title="AI Lead Scoring & Automation System",
    page_icon="💼",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for Glassmorphic/Modern Dark Mode appearance in Streamlit
st.markdown("""
<style>
    /* Main container background */
    .stApp {
        background-color: #0b0f19;
        color: #f8fafc;
    }
    
    /* Sidebar styling */
    section[data-testid="stSidebar"] {
        background-color: #111827 !important;
        border-right: 1px solid rgba(255, 255, 255, 0.05);
    }
    
    /* Metric styling */
    div[data-testid="metric-container"] {
        background-color: rgba(30, 41, 59, 0.45);
        border: 1px solid rgba(255, 255, 255, 0.08);
        padding: 15px;
        border-radius: 12px;
        box-shadow: 0 4px 15px rgba(0,0,0,0.2);
    }
    
    /* Header styling */
    .main-title {
        font-family: 'Outfit', sans-serif;
        font-size: 2.2rem;
        font-weight: 800;
        background: linear-gradient(to right, #60a5fa, #818cf8);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        margin-bottom: 5px;
    }
    
    .subtitle {
        color: #94a3b8;
        font-size: 1rem;
        margin-bottom: 25px;
        text-transform: uppercase;
        letter-spacing: 1px;
    }
</style>
""", unsafe_allow_html=True)


# --- BUSINESS SCORING ENGINE ---

def heuristic_scoring(note):
    """
    Evaluates customer note and assigns a score and classification.
    Implements rules from tieu_chi_cham_diem.txt.
    """
    if not isinstance(note, str) or not note.strip():
        return {
            "score": 0,
            "classification": "Rác",
            "reason": "Nội dung ghi chú trống.",
            "budget": "N/A",
            "property_type": "N/A",
            "location": "N/A",
            "urgency": "N/A"
        }

    score = 50
    reasons_plus = []
    reasons_minus = []
    
    extracted = {
        "budget": "Không rõ",
        "property_type": "Không rõ",
        "location": "Không rõ",
        "urgency": "Trung bình"
    }

    note_lower = note.lower()

    # --- 1. TIÊU CHÍ CỘNG 50 ĐIỂM (VIP) ---
    
    # Ngân sách >= 20 tỷ hoặc cụm từ khóa tài chính mạnh
    budget_match = re.search(r"(\d+)\s*(tỷ|ty)", note_lower)
    if budget_match:
        budget_val = int(budget_match.group(1))
        extracted["budget"] = f"{budget_val} tỷ"
        if budget_val >= 20:
            score += 50
            reasons_plus.append(f"Ngân sách lớn ({budget_val} tỷ >= 20 tỷ)")
    elif any(x in note_lower for x in ["tài chính mạnh", "tai chinh manh", "không thành vấn đề", "khong thanh van de", "tài chính cao", "tai chinh cao"]):
        score += 50
        extracted["budget"] = "Tài chính mạnh"
        reasons_plus.append("Có đề cập 'tài chính mạnh/không thành vấn đề'")

    # Loại hình cao cấp
    luxury_types = ["biệt thự đơn lập", "biet thu don lap", "biệt thự", "biet thu", "penthouse", "shophouse mặt đường lớn", "shophouse mat duong lon", "shophouse", "quỹ đất công nghiệp", "quy dat cong nghiep", "sàn văn phòng diện tích lớn", "san van phong dien tich lon", "sàn văn phòng", "san van phong"]
    for t in luxury_types:
        if t in note_lower:
            score += 50
            extracted["property_type"] = t.title()
            reasons_plus.append(f"Loại hình cao cấp ({t.title()})")
            break

    # Vị trí đắc địa
    prime_locations = ["quận 1", "quan 1", "ven sông", "ven song", "vinhomes ocean park", "ocean park", "phú mỹ hưng", "phu my hung"]
    for l in prime_locations:
        if l in note_lower:
            score += 50
            extracted["location"] = l.title()
            reasons_plus.append(f"Vị trí đắc địa ({l.title()})")
            break

    # Đối tượng khách hàng
    vip_clients = ["chủ doanh nghiệp", "chu doanh nghiep", "nhà đầu tư chuyên nghiệp", "nha dau tu chuyen nghiep", "mua sỉ", "mua si", "mua số lượng lớn", "mua so luong lon"]
    for c in vip_clients:
        if c in note_lower:
            score += 50
            reasons_plus.append(f"Đối tượng VIP ({c.capitalize()})")
            break

    # Tính cấp thiết & Minh bạch
    urgency_keywords = ["pháp lý chuẩn 100%", "phap ly chuan 100%", "sổ hồng riêng", "so hong rieng", "gặp trực tiếp chủ đầu tư để đàm phán", "gap truc tiep chu dau tu de dam phan", "gặp trực tiếp chủ đầu tư", "gap truc tiep chu dau tu"]
    for u in urgency_keywords:
        if u in note_lower:
            score += 50
            extracted["urgency"] = "Cao"
            reasons_plus.append("Cấp thiết & Minh bạch (Yêu cầu pháp lý/Muốn gặp trực tiếp CĐT)")
            break


    # --- 2. TIÊU CHÍ TRỪ 50 ĐIỂM (RÁC) ---
    
    # Yêu cầu phi thực tế (Ví dụ: Nhà Quận 1 giá 1-2 tỷ, nhà trung tâm bể bơi giá vài trăm triệu)
    if ("quận 1" in note_lower or "quan 1" in note_lower) and budget_match:
        budget_val = int(budget_match.group(1))
        if budget_val <= 3:
            score -= 50
            reasons_minus.append(f"Yêu cầu phi thực tế (Mua nhà Q1 với giá {budget_val} tỷ)")
    elif ("trung tâm" in note_lower or "trung tam" in note_lower or "sân vườn" in note_lower or "san vuon" in note_lower or "hồ bơi" in note_lower or "ho boi" in note_lower) and budget_match:
        budget_val = int(budget_match.group(1))
        if budget_val < 1:  # Dưới 1 tỷ
            score -= 50
            reasons_minus.append("Yêu cầu phi thực tế (Nhà trung tâm có sân vườn hồ bơi giá thấp vô lý)")
    elif "vài trăm triệu" in note_lower or "vai tram trieu" in note_lower:
        score -= 50
        reasons_minus.append("Yêu cầu phi thực tế (Giá quá thấp so với thị trường)")

    # Không có nhu cầu
    if any(x in note_lower for x in ["nhầm số", "nham so", "không có nhu cầu", "khong co nhu cau", "dữ liệu cũ", "du lieu cu", "nhầm ngành", "nham nganh"]):
        score -= 50
        reasons_minus.append("Không có nhu cầu (Báo nhầm số/ngành/dữ liệu cũ)")

    # Khách hàng không thiện chí
    if any(x in note_lower for x in ["hỏi giá cho vui", "hoi gia cho vui", "chưa có ý định mua", "chua co y dinh mua", "thái độ không hợp tác", "thai do khong hop tac"]):
        score -= 50
        reasons_minus.append("Khách hàng thiếu thiện chí (Hỏi chơi/Không hợp tác)")

    # Spam/Quảng cáo
    if any(x in note_lower for x in ["bảo hiểm", "bao hiem", "vay vốn", "vay von", "mời chào dịch vụ", "moi chao dich vu", "tuyển dụng", "tuyen dung"]):
        score -= 50
        reasons_minus.append("Spam/Quảng cáo (Chào mời bảo hiểm, vay vốn, tuyển dụng...)")

    # Thông tin liên lạc lỗi
    if any(x in note_lower for x in ["thuê bao", "thue bao", "gọi nhiều lần không bắt máy", "goi nhieu lan khong bat may", "không nghe máy", "khong nghe may", "không phản hồi zalo", "khong phan hoi zalo"]):
        score -= 50
        reasons_minus.append("Lỗi liên lạc (Thuê bao/Không nghe máy/Không rep Zalo)")


    # --- 3. ĐIỀU CHỈNH ĐIỂM SỐ & PHÂN LOẠI ---
    
    # Boundary correction
    if len(reasons_plus) > 0 and len(reasons_minus) == 0:
        score = min(score, 100)
    elif len(reasons_minus) > 0:
        score = max(score - 50, 0)
    else:
        score = 50  # Neutral baseline

    # Classification
    if score >= 90:
        classification = "VIP"
    elif score <= 20:
        classification = "Rác"
    else:
        classification = "Tiềm năng"

    # Assemble Vietnamese reason string
    reason_parts = []
    if reasons_plus:
        reason_parts.append("[CỘNG ĐIỂM: VIP] " + ", ".join(reasons_plus))
    if reasons_minus:
        reason_parts.append("[TRỪ ĐIỂM: RÁC] " + ", ".join(reasons_minus))
    
    if not reason_parts:
        if "chung cư" in note_lower or "chung cu" in note_lower or "nhà phố" in note_lower or "nha pho" in note_lower:
            reason_str = "Khách tìm mua chung cư/nhà phố trung cấp (3-10 tỷ) - Giữ điểm cơ sở."
        elif "vay" in note_lower or "ngân hàng" in note_lower or "ngan hang" in note_lower:
            reason_str = "Khách hàng cần vay ngân hàng, đang cân nhắc chính sách - Giữ điểm cơ sở."
        else:
            reason_str = "Nhu cầu thực tế, cần tư vấn thêm về pháp lý/vị trí - Giữ điểm cơ sở."
    else:
        reason_str = " | ".join(reason_parts)

    return {
        "score": score,
        "classification": classification,
        "reason": reason_str,
        "budget": extracted["budget"],
        "property_type": extracted["property_type"],
        "location": extracted["location"],
        "urgency": extracted["urgency"]
    }


def generate_excel_bytes(df_leads):
    """
    Generates a beautifully formatted openpyxl Excel file as a bytes stream.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "AI Lead Scoring Results"
    ws.views.sheetView[0].showGridLines = True
    
    # Font & Styles
    font_name = "Segoe UI"
    title_font = Font(name=font_name, size=15, bold=True, color="1E3A8A")
    subtitle_font = Font(name=font_name, size=10, italic=True, color="555555")
    header_font = Font(name=font_name, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Navy Accent
    
    cell_font = Font(name=font_name, size=10)
    bold_cell_font = Font(name=font_name, size=10, bold=True)
    
    fill_vip = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")     # Emerald Green Accent
    font_vip = Font(name=font_name, size=10, bold=True, color="065F46")
    
    fill_normal = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # Amber Gold Accent
    font_normal = Font(name=font_name, size=10, bold=True, color="92400E")
    
    fill_trash = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")   # Coral Red Accent
    font_trash = Font(name=font_name, size=10, bold=True, color="991B1B")

    center_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=False)
    
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    # Title Rows
    ws.merge_cells("A1:K1")
    ws["A1"] = "BÁO CÁO KHÁCH HÀNG TIỀM NĂNG - AI LEAD SCORING SYSTEM"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    
    ws.merge_cells("A2:K2")
    ws["A2"] = "Dữ liệu khách hàng đã chấm điểm tự động & kiểm duyệt thủ công (Human-in-the-loop)"
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    
    ws.append([]) # spacing row
    
    # Headers
    headers = [
        "Mã KH", "Họ tên", "Số điện thoại", "Nhu cầu chi tiết", 
        "Điểm số", "Phân loại", "Lý do chi tiết", 
        "Ngân sách", "Loại hình BĐS", "Vị trí yêu cầu", "Độ cấp thiết"
    ]
    ws.append(headers)
    header_row = 4
    
    # Format Headers
    for c_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=c_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    # Append Data from Dataframe
    for _, row in df_leads.iterrows():
        ws.append([
            row["Mã KH"],
            row["Họ tên"],
            row["Số điện thoại"],
            row["Nhu cầu chi tiết"],
            row["Điểm số"],
            row["Phân loại"],
            row["Lý do chi tiết"],
            row["Ngân sách (Trích xuất)"],
            row["Loại hình BĐS"],
            row["Vị trí yêu cầu"],
            row["Độ cấp thiết"]
        ])
        
    # Format Data Rows
    start_row = 5
    end_row = ws.max_row
    
    for r in range(start_row, end_row + 1):
        ws.row_dimensions[r].height = 20
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = cell_font
            cell.border = thin_border
            cell.alignment = left_align
            
        # Alignment Overrides
        ws.cell(row=r, column=1).alignment = center_align # ID
        ws.cell(row=r, column=3).alignment = center_align # SĐT
        ws.cell(row=r, column=5).alignment = center_align # Score
        ws.cell(row=r, column=6).alignment = center_align # Classification
        ws.cell(row=r, column=11).alignment = center_align # Urgency
        
        # Classification Conditional Colors
        score_cell = ws.cell(row=r, column=5)
        score_cell.font = bold_cell_font
        
        class_cell = ws.cell(row=r, column=6)
        val = class_cell.value
        if val == "VIP":
            class_cell.fill = fill_vip
            class_cell.font = font_vip
        elif val == "Tiềm năng":
            class_cell.fill = fill_normal
            class_cell.font = font_normal
        elif val == "Rác":
            class_cell.fill = fill_trash
            class_cell.font = font_trash
            
    # Auto-adjust Column Widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        for idx, cell in enumerate(col):
            if idx < 3: # Skip title blocks
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
                
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Override wide columns manually for readability
    ws.column_dimensions['D'].width = 40  # Nhu cầu
    ws.column_dimensions['G'].width = 50  # Lý do
    
    ws.row_dimensions[4].height = 26
    
    # Save to buffer
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer.getvalue()


# --- STREAMLIT PAGE FLOW ---

st.markdown("<div class='main-title'>AI LEAD SCORING & AUTOMATION SYSTEM</div>", unsafe_allow_html=True)
st.markdown("<div class='subtitle'>Platform Bất Động Sản Tự Động Hóa • Streamlit Interface</div>", unsafe_allow_html=True)

# 1. Sidebar Config
st.sidebar.markdown("### 🛠️ CẤU HÌNH LIÊN KẾT")
sheet_url = st.sidebar.text_input(
    "Đường dẫn Google Sheets:",
    value="https://docs.google.com/spreadsheets/d/1hRvHE6RXm1peVG07avfApPEHocOcPld9IA94hE3vUGE/edit?gid=0#gid=0"
)

# Initialize Session State for Dataframe
if "df_leads" not in st.session_state:
    st.session_state.df_leads = None
    
# Sync Google Sheets Button
if st.sidebar.button("🔄 ĐỒNG BỘ DỮ LIỆU"):
    # Convert edit/sharing link to export link
    export_url = sheet_url
    if "/edit" in export_url:
        export_url = export_url.split("/edit")[0] + "/export?format=csv"
        gid_match = re.search(r"gid=(\d+)", sheet_url)
        if gid_match:
            export_url += f"&gid={gid_match.group(1)}"
            
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }
    
    try:
        with st.spinner("Đang tải dữ liệu từ Google Sheets..."):
            response = requests.get(export_url, headers=headers, timeout=10)
            if response.status_code == 200:
                df = pd.read_csv(requests.compat.StringIO(response.text))
                st.sidebar.success("Tải thành công từ Google Sheets!")
            else:
                st.sidebar.warning(f"Lỗi tải Sheets ({response.status_code}). Sử dụng dữ liệu mẫu.")
                df = pd.read_csv("mock_leads.csv")
    except Exception as e:
        st.sidebar.warning(f"Không kết nối được Sheets. Đang sử dụng dữ liệu mẫu cuc bộ. (Lỗi: {e})")
        df = pd.read_csv("mock_leads.csv")
        
    # Map headers
    col_name = next((c for c in df.columns if any(x in str(c).lower() for x in ['tên', 'ho ten', 'name', 'khách hàng']) and not any(y in str(c).lower() for y in ['nhu cầu', 'ghi chú', 'note'])), df.columns[1] if len(df.columns) > 1 else df.columns[0])
    col_phone = next((c for c in df.columns if any(x in str(c).lower() for x in ['thoại', 'sđt', 'sdt', 'phone', 'liên hệ'])), df.columns[2] if len(df.columns) > 2 else df.columns[0])
    col_note = next((c for c in df.columns if any(x in str(c).lower() for x in ['nhu cầu', 'ghi chú', 'note', 'inquiry', 'mô tả'])), df.columns[3] if len(df.columns) > 3 else df.columns[0])
    
    # Process initial scoring
    leads_list = []
    for idx, row in df.iterrows():
        note_val = str(row[col_note]) if pd.notna(row[col_note]) else ""
        res = heuristic_scoring(note_val)
        leads_list.append({
            "Mã KH": idx + 1,
            "Họ tên": str(row[col_name]) if pd.notna(row[col_name]) else "Không tên",
            "Số điện thoại": str(row[col_phone]) if pd.notna(row[col_phone]) else "Không có",
            "Nhu cầu chi tiết": note_val,
            "Điểm số": res["score"],
            "Phân loại": res["classification"],
            "Lý do chi tiết": res["reason"],
            "Ngân sách (Trích xuất)": res["budget"],
            "Loại hình BĐS": res["property_type"],
            "Vị trí yêu cầu": res["location"],
            "Độ cấp thiết": res["urgency"]
        })
        
    st.session_state.df_leads = pd.DataFrame(leads_list)
    st.sidebar.success(f"Đã nạp {len(st.session_state.df_leads)} leads!")

# Load Mock Data initially if session state is empty
if st.session_state.df_leads is None:
    if os.path.exists("mock_leads.csv"):
        df = pd.read_csv("mock_leads.csv")
        leads_list = []
        for idx, row in df.iterrows():
            note_val = str(row['Nhu cầu khách hàng']) if pd.notna(row['Nhu cầu khách hàng']) else ""
            res = heuristic_scoring(note_val)
            leads_list.append({
                "Mã KH": idx + 1,
                "Họ tên": str(row['Họ tên']),
                "Số điện thoại": str(row['Số điện thoại']),
                "Nhu cầu chi tiết": note_val,
                "Điểm số": res["score"],
                "Phân loại": res["classification"],
                "Lý do chi tiết": res["reason"],
                "Ngân sách (Trích xuất)": res["budget"],
                "Loại hình BĐS": res["property_type"],
                "Vị trí yêu cầu": res["location"],
                "Độ cấp thiết": res["urgency"]
            })
        st.session_state.df_leads = pd.DataFrame(leads_list)
    else:
        st.error("Không tìm thấy tệp mock_leads.csv trong thư mục!")

# If we have data, render layout
if st.session_state.df_leads is not None:
    df_active = st.session_state.df_leads
    
    # 2. Main KPI Metrics
    st.markdown("### 📊 THÔNG KÊ PHÂN TÍCH CHẤT LƯỢNG")
    col_kpi1, col_kpi2, col_kpi3, col_kpi4 = st.columns(4)
    
    total_count = len(df_active)
    vip_count = len(df_active[df_active["Phân loại"] == "VIP"])
    potential_count = len(df_active[df_active["Phân loại"] == "Tiềm năng"])
    trash_count = len(df_active[df_active["Phân loại"] == "Rác"])
    
    with col_kpi1:
        st.metric("Tổng Số Khách Hàng", f"{total_count} leads")
    with col_kpi2:
        st.metric("Khách VIP (Siêu Tiềm Năng)", f"{vip_count} leads", delta=f"{int(vip_count/total_count*100) if total_count>0 else 0}%")
    with col_kpi3:
        st.metric("Khách Tiềm Năng (Tầm Trung)", f"{potential_count} leads", delta=f"{int(potential_count/total_count*100) if total_count>0 else 0}%")
    with col_kpi4:
        st.metric("Khách Hàng Rác (Không mua)", f"{trash_count} leads", delta=f"-{int(trash_count/total_count*100) if total_count>0 else 0}%", delta_color="inverse")
        
    st.markdown("---")
    
    # 3. Interactive Editor (Human-in-the-loop)
    st.markdown("### ✏️ KIỂM DUYỆT & CHỈNH SỬA THỦ CÔNG (HUMAN-IN-THE-LOOP)")
    st.info("💡 Bạn có thể chỉnh sửa trực tiếp điểm số, phân loại, thông tin nhu cầu hoặc lý do đánh giá ngay trên bảng dưới đây trước khi xuất Excel.")
    
    # Configure columns for editing
    edited_df = st.data_editor(
        df_active,
        column_config={
            "Mã KH": st.column_config.NumberColumn(disabled=True),
            "Họ tên": st.column_config.TextColumn(width="medium"),
            "Số điện thoại": st.column_config.TextColumn(width="small"),
            "Nhu cầu chi tiết": st.column_config.TextColumn(width="large"),
            "Điểm số": st.column_config.NumberColumn(min_value=0, max_value=100, step=5),
            "Phân loại": st.column_config.SelectboxColumn(options=["VIP", "Tiềm năng", "Rác"]),
            "Lý do chi tiết": st.column_config.TextColumn(width="large"),
            "Ngân sách (Trích xuất)": st.column_config.TextColumn(),
            "Loại hình BĐS": st.column_config.TextColumn(),
            "Vị trí yêu cầu": st.column_config.TextColumn(),
            "Độ cấp thiết": st.column_config.SelectboxColumn(options=["Thấp", "Trung bình", "Cao"])
        },
        hide_index=True,
        num_rows="dynamic",
        use_container_width=True
    )
    
    # Save the edited edits back into st.session_state
    st.session_state.df_leads = edited_df
    
    # 4. Action Buttons (Export to Excel)
    st.markdown("### 📤 KẾT XUẤT DỮ LIỆU BÀN GIAO")
    
    excel_data = generate_excel_bytes(edited_df)
    
    col_btn1, col_btn2 = st.columns([1, 4])
    with col_btn1:
        st.download_button(
            label="📥 Tải Báo Cáo Excel (.xlsx)",
            data=excel_data,
            file_name="ket_qua_cham_diem_lead_streamlit.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    with col_btn2:
        st.success("Tệp Excel đã sẵn sàng để tải xuống. Được định dạng chuyên nghiệp với các cột chi tiết và tô màu phân loại tự động!")

    # 5. Charts / Visuals
    st.markdown("### 📈 BIỂU ĐỒ PHÂN BỐ KHÁCH HÀNG")
    col_chart1, col_chart2 = st.columns(2)
    
    with col_chart1:
        # Simple distribution bar chart
        chart_data = pd.DataFrame({
            "Phân loại": ["VIP", "Tiềm năng", "Rác"],
            "Số lượng": [vip_count, potential_count, trash_count]
        })
        st.bar_chart(chart_data.set_index("Phân loại"))
        
    with col_chart2:
        st.write("📊 **Chi tiết cơ sở dữ liệu sau chỉnh sửa:**")
        st.dataframe(
            edited_df[["Họ tên", "Số điện thoại", "Điểm số", "Phân loại", "Lý do chi tiết"]],
            use_container_width=True,
            hide_index=True
        )
