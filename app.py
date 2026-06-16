import os
import sys
import io
import json
import re

# Force UTF-8 encoding for Windows console logs
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import pandas as pd
import requests
from flask import Flask, jsonify, request, render_template, send_file
import google.generativeai as genai
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

app = Flask(__name__, template_folder='templates')

# In-memory storage for leads
leads_db = []

# Default google sheet url for convenience
DEFAULT_SHEET_URL = "https://docs.google.com/spreadsheets/d/1hRvHE6RXm1peVG07avfApPEHocOcPld9IA94hE3vUGE/edit?gid=0#gid=0"

def init_mock_data():
    """Load mock data from mock_leads.csv if it exists, otherwise create it"""
    global leads_db
    mock_file = os.path.join(os.path.dirname(__file__), 'mock_leads.csv')
    if os.path.exists(mock_file):
        try:
            df = pd.read_csv(mock_file)
            leads_db = []
            for _, row in df.iterrows():
                leads_db.append({
                    "id": int(row['ID']),
                    "name": str(row['Họ tên']),
                    "phone": str(row['Số điện thoại']),
                    "note": str(row['Nhu cầu khách hàng']),
                    "score": 50,  # Default baseline score
                    "classification": "Chưa đánh giá",
                    "reason": "Chưa chạy AI chấm điểm.",
                    "extracted_info": {
                        "budget": "N/A",
                        "property_type": "N/A",
                        "location": "N/A",
                        "urgency": "N/A"
                    },
                    "human_reviewed": False
                })
            print(f"Loaded {len(leads_db)} leads from mock_leads.csv")
        except Exception as e:
            print("Error loading mock_leads.csv:", e)

# Initialize database with mock data
init_mock_data()


def fetch_from_google_sheet(sheet_url):
    """
    Fetches data from a Google Sheet URL by transforming it into a CSV download link.
    """
    # Convert edit/sharing link to export link
    export_url = sheet_url
    if "/edit" in export_url:
        export_url = export_url.split("/edit")[0] + "/export?format=csv"
        # Extract gid if present
        gid_match = re.search(r"gid=(\d+)", sheet_url)
        if gid_match:
            export_url += f"&gid={gid_match.group(1)}"
    else:
        # Fallback default export if no edit is found
        export_url = "https://docs.google.com/spreadsheets/d/1hRvHE6RXm1peVG07avfApPEHocOcPld9IA94hE3vUGE/export?format=csv&gid=0"

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }
    
    response = requests.get(export_url, headers=headers, timeout=10)
    if response.status_code != 200:
        raise Exception(f"Google Sheet returned status code {response.status_code}")
        
    df = pd.read_csv(requests.compat.StringIO(response.text))
    return df


def simulate_ai_scoring(note):
    """
    A fallback heuristic rule-based scorer that mimics the logic of tieu_chi_cham_diem.txt.
    Used when Gemini API key is not provided.
    """
    score = 50
    reasons_plus = []
    reasons_minus = []
    
    extracted = {
        "budget": "Không rõ",
        "property_type": "Không rõ",
        "location": "Không rõ",
        "urgency": "Trung bình"
    }

    note_lower = note.lower()

    # Rule 1: VIP (+50)
    # Budget >= 20B or strong finance
    budget_match = re.search(r"(\d+)\s*(tỷ|ty)", note_lower)
    if budget_match:
        budget_val = int(budget_match.group(1))
        extracted["budget"] = f"{budget_val} tỷ"
        if budget_val >= 20:
            score += 50
            reasons_plus.append(f"Ngân sách lớn ({budget_val} tỷ >= 20 tỷ)")
    elif "tài chính mạnh" in note_lower or "không thành vấn đề" in note_lower or "tài chính cao" in note_lower:
        score += 50
        extracted["budget"] = "Tài chính mạnh"
        reasons_plus.append("Có đề cập 'tài chính mạnh/không thành vấn đề'")

    # Luxury property types
    luxury_types = ["biệt thự", "biet thu", "penthouse", "shophouse", "quỹ đất công nghiệp", "quy dat cong nghiep", "sàn văn phòng", "san van phong"]
    for t in luxury_types:
        if t in note_lower:
            score += 50
            extracted["property_type"] = t.capitalize()
            reasons_plus.append(f"Loại hình cao cấp ({t.capitalize()})")
            break

    # Prime locations
    prime_locs = ["quận 1", "quan 1", "ven sông", "ven song", "vinhomes ocean park", "ocean park", "phú mỹ hưng", "phu my hung"]
    for l in prime_locs:
        if l in note_lower:
            score += 50
            extracted["location"] = l.capitalize()
            reasons_plus.append(f"Vị trí đắc địa ({l.capitalize()})")
            break

    # Client type
    if any(x in note_lower for x in ["chủ doanh nghiệp", "chu doanh nghiep", "nhà đầu tư chuyên nghiệp", "nha dau tu chuyên nghiep", "mua sỉ", "mua si", "mua số lượng lớn", "mua so luong lon"]):
        score += 50
        reasons_plus.append("Đối tượng khách hàng tiềm năng (Chủ DN/NĐT/Mua sỉ)")

    # Urgency & transparency
    if any(x in note_lower for x in ["pháp lý chuẩn", "sổ hồng riêng", "so hong rieng", "gặp trực tiếp chủ đầu tư", "gap truc tiep chu dau tu"]):
        score += 50
        extracted["urgency"] = "Cao (Yêu cầu pháp lý/Gặp trực tiếp)"
        reasons_plus.append("Tính cấp thiết & Minh bạch pháp lý")

    # Rule 2: Trash (-50)
    # Unrealistic pricing (e.g. Q1 price < 5B)
    if ("quận 1" in note_lower or "quan 1" in note_lower) and budget_match:
        budget_val = int(budget_match.group(1))
        if budget_val <= 3:
            score -= 50
            reasons_minus.append(f"Yêu cầu phi thực tế (Nhà Q1 giá {budget_val} tỷ)")
    elif "hồ bơi" in note_lower and budget_match:
        budget_val = int(budget_match.group(1))
        if budget_val < 1:
            score -= 50
            reasons_minus.append("Yêu cầu phi thực tế (Nhà trung tâm bể bơi giá quá thấp)")

    # No demand
    if any(x in note_lower for x in ["nhầm số", "nham so", "không có nhu cầu", "khong co nhu cau", "dữ liệu cũ", "du lieu cu", "nhầm ngành", "nham nganh"]):
        score -= 50
        reasons_minus.append("Khách hàng báo nhầm số hoặc không có nhu cầu")

    # Uncooperative
    if any(x in note_lower for x in ["hỏi giá cho vui", "hoi gia cho vui", "chưa có ý định mua", "chua co y dinh mua", "không hợp tác", "khong hop tac"]):
        score -= 50
        reasons_minus.append("Khách hàng hỏi chơi, không thiện chí hợp tác")

    # Spam/Ads
    if any(x in note_lower for x in ["bảo hiểm", "bao hiem", "vay vốn", "vay von", "mời chào", "moi chao", "tuyển dụng"]):
        score -= 50
        reasons_minus.append("Nội dung spam hoặc quảng cáo dịch vụ khác (bảo hiểm, vay vốn...)")

    # Communication errors
    if any(x in note_lower for x in ["thuê bao", "thue bao", "không bắt máy", "khong bat may", "không phản hồi zalo", "khong phan hoi zalo"]):
        score -= 50
        reasons_minus.append("Lỗi liên lạc (Thuê bao, không nghe máy, không check Zalo)")

    # Adjust score boundaries
    if len(reasons_plus) > 0 and len(reasons_minus) == 0:
        score = min(score, 100)
    elif len(reasons_minus) > 0:
        score = max(score - 50, 0)
    else:
        score = 50  # Neutral baseline

    # Classify
    if score >= 90:
        classification = "VIP"
    elif score <= 20:
        classification = "Rác"
    else:
        classification = "Tiềm năng"

    # Compile reasons
    reason_str = ""
    if reasons_plus:
        reason_str += "[CỘNG ĐIỂM] " + "; ".join(reasons_plus) + ". "
    if reasons_minus:
        reason_str += "[TRỪ ĐIỂM] " + "; ".join(reasons_minus) + ". "
    if not reasons_plus and not reasons_minus:
        reason_str = "Khách hàng tầm trung, nhu cầu thực tế hoặc chưa đủ dữ liệu để xếp loại đặc biệt."

    return {
        "score": score,
        "classification": classification,
        "reason": reason_str,
        "extracted_info": extracted
    }


def call_gemini_scoring(note, api_key):
    """
    Calls the Gemini API using the provided key to score the lead note.
    """
    try:
        genai.configure(api_key=api_key)
        # Using gemini-1.5-flash for speed and reliability
        model = genai.GenerativeModel('gemini-1.5-flash')
        
        system_instructions = (
            "Bạn là chuyên gia chấm điểm khách hàng tiềm năng ngành Bất động sản theo các quy tắc sau:\n"
            "Mỗi khách hàng xuất phát với điểm cơ sở là 50 điểm.\n"
            "1. Cộng 50 điểm (VIP) nếu phát hiện các yếu tố: Ngân sách lớn (>=20 tỷ hoặc 'tài chính mạnh'), "
            "Loại hình cao cấp (Biệt thự, Penthouse, Shophouse đường lớn, Quỹ đất CN, Sàn VP lớn), "
            "Vị trí đắc địa (Q1, ven sông, Vinhomes Ocean Park, Phú Mỹ Hưng), "
            "Đối tượng VIP (Chủ doanh nghiệp, NĐT chuyên nghiệp, mua sỉ), "
            "Tính cấp thiết (Pháp lý chuẩn 100%, sổ hồng riêng, muốn gặp trực tiếp CĐT).\n"
            "2. Trừ 50 điểm (Rác) nếu phát hiện: Đòi mua giá siêu rẻ phi thực tế (Q1 giá 1-2 tỷ, trung tâm bể bơi giá vài trăm triệu), "
            "Không có nhu cầu (Nhầm số, không có nhu cầu, dữ liệu cũ), "
            "Không thiện chí (Hỏi cho vui, thái độ không hợp tác), "
            "Spam/Quảng cáo (Mời bảo hiểm, mời vay vốn), "
            "Lỗi liên lạc (Thuê bao, không nghe máy, không phản hồi Zalo).\n"
            "3. Giữ nguyên 50 điểm hoặc cộng nhẹ (Tiềm năng) cho các trường hợp khác: Mua chung cư/nhà phố 3-10 tỷ, cần vay ngân hàng, có nhu cầu thực cần tư vấn thêm.\n"
            "Điểm số cuối cùng phải giới hạn trong khoảng [0, 100].\n"
            "Hãy phân tích và trả về đúng một chuỗi JSON có cấu trúc như sau (không kèm markdown block ```json):\n"
            "{\n"
            '  "score": <số nguyên từ 0 đến 100>,\n'
            '  "classification": "<VIP | Tiềm năng | Rác>",\n'
            '  "reason": "<Lý do chi tiết tiếng Việt>",\n'
            '  "extracted_info": {\n'
            '    "budget": "<Ngân sách trích xuất>",\n'
            '    "property_type": "<Loại hình BĐS>",\n'
            '    "location": "<Vị trí yêu cầu>",\n'
            '    "urgency": "<Mức độ cấp thiết>"\n'
            "  }\n"
            "}"
        )
        
        prompt = f"Phân tích nhu cầu của khách hàng sau và chấm điểm:\n\n{note}"
        
        response = model.generate_content(
            contents=prompt,
            generation_config={"response_mime_type": "application/json"},
            safety_settings=[],
            tools=[]
        )
        
        raw_text = response.text.strip()
        # Clean any markdown syntax if present
        raw_text = re.sub(r"^```json\s*", "", raw_text)
        raw_text = re.sub(r"\s*```$", "", raw_text)
        
        result = json.loads(raw_text)
        return result
    except Exception as e:
        print("Gemini API call failed, falling back to simulated engine. Error:", e)
        # Fallback to simulation
        return simulate_ai_scoring(note)


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/leads', methods=['GET'])
def get_leads():
    return jsonify(leads_db)


@app.route('/api/fetch', methods=['POST'])
def fetch_leads():
    global leads_db
    data = request.json or {}
    sheet_url = data.get('url', '').strip()
    
    if not sheet_url:
        # Load mock data if no URL is provided
        init_mock_data()
        return jsonify({"status": "success", "message": "Đã tải dữ liệu mẫu thành công!", "count": len(leads_db)})
        
    try:
        df = fetch_from_google_sheet(sheet_url)
        
        # Verify columns mapping
        # Expecting something like: Họ tên, Số điện thoại, Nhu cầu khách hàng (or similar)
        # Let's inspect columns and find the closest match
        col_name = None
        col_phone = None
        col_note = None
        
        for col in df.columns:
            col_lower = str(col).lower()
            if 'tên' in col_lower or 'ho ten' in col_lower or 'name' in col_lower or 'khách hàng' in col_lower:
                col_name = col
            if 'thoại' in col_lower or 'sđt' in col_lower or 'sdt' in col_lower or 'phone' in col_lower or 'liên hệ' in col_lower:
                col_phone = col
            if 'nhu cầu' in col_lower or 'ghi chú' in col_lower or 'note' in col_lower or 'inquiry' in col_lower or 'mô tả' in col_lower:
                col_note = col

        # Direct fallbacks if headers don't match
        if not col_name: col_name = df.columns[1] if len(df.columns) > 1 else df.columns[0]
        if not col_phone: col_phone = df.columns[2] if len(df.columns) > 2 else df.columns[0]
        if not col_note: col_note = df.columns[3] if len(df.columns) > 3 else df.columns[0]
        
        leads_db = []
        for idx, row in df.iterrows():
            leads_db.append({
                "id": idx + 1,
                "name": str(row[col_name]) if pd.notna(row[col_name]) else "Không tên",
                "phone": str(row[col_phone]) if pd.notna(row[col_phone]) else "Không có",
                "note": str(row[col_note]) if pd.notna(row[col_note]) else "",
                "score": 50,
                "classification": "Chưa đánh giá",
                "reason": "Chưa chạy AI chấm điểm.",
                "extracted_info": {
                    "budget": "N/A",
                    "property_type": "N/A",
                    "location": "N/A",
                    "urgency": "N/A"
                },
                "human_reviewed": False
            })
            
        return jsonify({"status": "success", "message": "Đã đồng bộ dữ liệu từ Google Sheets thành công!", "count": len(leads_db)})
        
    except Exception as e:
        print("Error fetching sheet:", e)
        return jsonify({
            "status": "error", 
            "message": f"Không thể tải dữ liệu từ Google Sheets (Lỗi: {str(e)}). Hệ thống sẽ tự động dùng dữ liệu mẫu thay thế."
        }), 400


@app.route('/api/score', methods=['POST'])
def score_leads():
    global leads_db
    data = request.json or {}
    user_api_key = data.get('api_key', '').strip()
    
    # Priority order for API key:
    # 1. User provided API key from client UI
    # 2. Server-side environment variable GEMINI_API_KEY
    api_key = user_api_key or os.environ.get("GEMINI_API_KEY")
    
    if not leads_db:
        return jsonify({"status": "error", "message": "Không có dữ liệu khách hàng để chấm điểm. Hãy tải dữ liệu trước!"}), 400
        
    scored_count = 0
    mode = "Gemini AI" if api_key else "Simulated AI (Rule-based Engine)"
    
    for lead in leads_db:
        # Skip if already reviewed by human, or optionally allow re-scoring
        if lead.get('human_reviewed', False):
            continue
            
        note = lead['note']
        if not note:
            lead['score'] = 0
            lead['classification'] = "Rác"
            lead['reason'] = "Nội dung nhu cầu trống."
            continue
            
        if api_key:
            res = call_gemini_scoring(note, api_key)
        else:
            res = simulate_ai_scoring(note)
            
        lead['score'] = int(res.get('score', 50))
        lead['classification'] = res.get('classification', 'Tiềm năng')
        lead['reason'] = res.get('reason', '')
        lead['extracted_info'] = res.get('extracted_info', {})
        scored_count += 1
        
    return jsonify({
        "status": "success", 
        "message": f"Chấm điểm thành công {scored_count} khách hàng bằng {mode}!", 
        "leads": leads_db,
        "mode": mode
    })


@app.route('/api/leads/<int:lead_id>', methods=['PUT'])
def update_lead(lead_id):
    global leads_db
    data = request.json or {}
    
    for lead in leads_db:
        if lead['id'] == lead_id:
            lead['score'] = int(data.get('score', lead['score']))
            lead['classification'] = data.get('classification', lead['classification'])
            lead['reason'] = data.get('reason', lead['reason'])
            lead['human_reviewed'] = True
            return jsonify({"status": "success", "lead": lead})
            
    return jsonify({"status": "error", "message": "Không tìm thấy khách hàng!"}), 404


@app.route('/api/export', methods=['POST'])
def export_leads():
    global leads_db
    if not leads_db:
        return jsonify({"status": "error", "message": "Không có dữ liệu để xuất file!"}), 400
        
    try:
        # Convert leads data to flat list
        flat_data = []
        for lead in leads_db:
            flat_data.append({
                "Mã KH": lead['id'],
                "Họ tên": lead['name'],
                "Số điện thoại": lead['phone'],
                "Mô tả Nhu cầu": lead['note'],
                "Điểm tiềm năng (AI)": lead['score'],
                "Phân loại": lead['classification'],
                "Lý do chi tiết": lead['reason'],
                "Ngân sách (Trích xuất)": lead['extracted_info'].get('budget', 'N/A'),
                "Loại hình BĐS": lead['extracted_info'].get('property_type', 'N/A'),
                "Vị trí yêu cầu": lead['extracted_info'].get('location', 'N/A'),
                "Mức độ cấp thiết": lead['extracted_info'].get('urgency', 'N/A'),
                "Kiểm duyệt bởi": "Người dùng (Human)" if lead['human_reviewed'] else "AI tự động"
            })
            
        df = pd.DataFrame(flat_data)
        
        # Output excel file path
        output_dir = os.path.dirname(__file__)
        file_path = os.path.join(output_dir, "danh_sach_khach_hang_tiem_nang.xlsx")
        
        # Write and Format Excel using openpyxl
        wb = Workbook()
        ws = wb.active
        ws.title = "Lead Scoring Results"
        ws.views.sheetView[0].showGridLines = True
        
        # Design system styles (Navy Accent theme)
        font_name = "Segoe UI"
        header_font = Font(name=font_name, size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Dark Navy
        
        title_font = Font(name=font_name, size=16, bold=True, color="1E3A8A")
        subtitle_font = Font(name=font_name, size=10, italic=True, color="555555")
        
        cell_font = Font(name=font_name, size=10)
        bold_cell_font = Font(name=font_name, size=10, bold=True)
        
        # Classification fills
        fill_vip = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")     # Light green
        font_vip = Font(name=font_name, size=10, bold=True, color="065F46")
        
        fill_normal = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # Light yellow
        font_normal = Font(name=font_name, size=10, bold=True, color="92400E")
        
        fill_trash = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")   # Light red
        font_trash = Font(name=font_name, size=10, bold=True, color="991B1B")

        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        
        thin_side = Side(style='thin', color='DDDDDD')
        border_cell = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        # Write custom header block
        ws.merge_cells("A1:L1")
        ws["A1"] = "BÁO CÁO KHÁCH HÀNG TIỀM NĂNG - AI LEAD SCORING SYSTEM"
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        
        ws.merge_cells("A2:L2")
        ws["A2"] = "Hệ thống tự động chấm điểm khách hàng & duyệt Human-in-the-loop"
        ws["A2"].font = subtitle_font
        ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
        
        # Empty row
        ws.append([])
        
        # Add DataFrame headers
        headers_row = list(df.columns)
        ws.append(headers_row)
        header_row_num = 4
        
        # Format Headers
        for col_idx in range(1, len(headers_row) + 1):
            cell = ws.cell(row=header_row_num, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            
        # Append data rows
        for _, row in df.iterrows():
            row_data = list(row)
            ws.append(row_data)
            
        # Format Data Rows
        start_row = 5
        end_row = ws.max_row
        
        for r in range(start_row, end_row + 1):
            ws.row_dimensions[r].height = 20
            
            # Format cells
            for c in range(1, len(headers_row) + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = cell_font
                cell.border = border_cell
                cell.alignment = left_align
                
            # Alignment override
            ws.cell(row=r, column=1).alignment = center_align # ID
            ws.cell(row=r, column=3).alignment = center_align # SDT
            ws.cell(row=r, column=5).alignment = center_align # Score
            ws.cell(row=r, column=6).alignment = center_align # Classification
            ws.cell(row=r, column=12).alignment = center_align # Reviewed by
            
            # Custom styling for Score & Classification
            score_cell = ws.cell(row=r, column=5)
            score_cell.font = bold_cell_font
            
            class_cell = ws.cell(row=r, column=6)
            val = class_cell.value
            if val == "VIP":
                class_cell.fill = fill_vip
                class_cell.font = font_vip
            elif val == "Tiềm năng":
                class_cell.fill = fill_normal
                class_cell.font = font_normal
            elif val == "Rác":
                class_cell.fill = fill_trash
                class_cell.font = font_trash

        # Auto-fit columns
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            
            # Find the longest text in the column (skipping the title rows)
            for idx, cell in enumerate(col):
                if idx < 3: # skip titles
                    continue
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        # Specific column manual adjustments for safety
        ws.column_dimensions['D'].width = 40  # Nhu cầu
        ws.column_dimensions['G'].width = 50  # Lý do
        
        # Set height for header
        ws.row_dimensions[4].height = 26
        
        # Save workbook
        wb.save(file_path)
        
        return send_file(file_path, as_attachment=True, download_name="danh_sach_khach_hang_tiem_nang.xlsx")
        
    except Exception as e:
        print("Error exporting Excel:", e)
        return jsonify({"status": "error", "message": f"Lỗi xuất Excel: {str(e)}"}), 500


if __name__ == '__main__':
    # Print welcome status
    print("--------------------------------------------------")
    print("Khoi dong server AI Lead Scoring & Automation...")
    print("Giao dien Web App chay tai: http://localhost:5000")
    print("--------------------------------------------------")
    app.run(host='0.0.0.0', port=5000, debug=True)
