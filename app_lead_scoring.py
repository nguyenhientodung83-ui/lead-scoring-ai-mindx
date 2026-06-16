import os
import sys
import io
import re
import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Force UTF-8 encoding for Windows console to prevent UnicodeEncodeError
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# Constants
SHEET_URL = "https://docs.google.com/spreadsheets/d/1hRvHE6RXm1peVG07avfApPEHocOcPld9IA94hE3vUGE/export?format=csv&gid=0"
OUTPUT_FILE = "ket_qua_cham_diem_lead.xlsx"
FALLBACK_FILE = "mock_leads.csv"


def fetch_data():
    """
    Downloads customer data from the Google Sheet link.
    If it fails due to network, 404 (private sheet), or other errors,
    it will look for a local CSV/Excel file or fallback to mock_leads.csv.
    """
    print("==================================================")
    print("1. DONG BO DU LIEU KHACH HANG")
    print("==================================================")
    
    # Step 1: Try to download from Google Sheet
    print(f"Dang tai du lieu tu Google Sheets URL...")
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }
    try:
        response = requests.get(SHEET_URL, headers=headers, timeout=10)
        if response.status_code == 200:
            print("-> Tai thanh cong tu Google Sheets!")
            df = pd.read_csv(requests.compat.StringIO(response.text))
            return df
        else:
            print(f"-> Google Sheets tra ve ma loi: {response.status_code} (co the tệp dang o che do rieng tu).")
    except Exception as e:
        print(f"-> Khong the ket noi den Google Sheets (Loi: {e})")

    # Step 2: Try to find local xlsx or csv files in the workspace (excluding output files)
    print("Dang tim kiem file du lieu cuc bo trong thu muc...")
    current_dir = os.path.dirname(os.path.abspath(__file__)) if __file__ else os.getcwd()
    for file in os.listdir(current_dir):
        file_lower = file.lower()
        if file_lower.endswith(('.xlsx', '.csv')):
            # Skip output/mock files
            if file in [OUTPUT_FILE, "danh_sach_khach_hang_tiem_nang.xlsx", "requirements.txt"]:
                continue
            if file == FALLBACK_FILE:
                continue
            
            file_path = os.path.join(current_dir, file)
            print(f"-> Tim thay file du lieu cuc bo: {file}")
            try:
                if file_lower.endswith('.xlsx'):
                    return pd.read_excel(file_path)
                else:
                    return pd.read_csv(file_path)
            except Exception as ex:
                print(f"-> Loi khi doc file {file}: {ex}")

    # Step 3: Fallback to mock_leads.csv
    print(f"Khong tim thay du lieu moi. Tu dong nap file du lieu mau: {FALLBACK_FILE}")
    fallback_path = os.path.join(current_dir, FALLBACK_FILE)
    if os.path.exists(fallback_path):
        try:
            return pd.read_csv(fallback_path)
        except Exception as e:
            print(f"-> Loi khi doc file du lieu mau: {e}")
            sys.exit("Khong the load bat ky nguon du lieu nao. Script dung hoat dong.")
    else:
        sys.exit(f"Khong tim thay file {FALLBACK_FILE} de lam du lieu du phong.")


def heuristic_scoring(note):
    """
    Evaluates customer note and assigns a score and classification.
    Implements rules from tieu_chi_cham_diem.txt.
    """
    if not isinstance(note, str) or not note.strip():
        return {
            "score": 0,
            "classification": "Rác",
            "reason": "Noi dung ghi chu trong.",
            "budget": "N/A",
            "property_type": "N/A",
            "location": "N/A",
            "urgency": "N/A"
        }

    score = 50
    reasons_plus = []
    reasons_minus = []
    
    extracted = {
        "budget": "Khong ro",
        "property_type": "Khong ro",
        "location": "Khong ro",
        "urgency": "Trung binh"
    }

    note_lower = note.lower()

    # --- 1. TIÊU CHÍ CỘNG 50 ĐIỂM (VIP) ---
    
    # Ngân sách >= 20 tỷ hoặc cụm từ khóa tài chính mạnh
    budget_match = re.search(r"(\d+)\s*(tỷ|ty)", note_lower)
    if budget_match:
        budget_val = int(budget_match.group(1))
        extracted["budget"] = f"{budget_val} tỷ"
        if budget_val >= 20:
            score += 50
            reasons_plus.append(f"Ngan sach lon ({budget_val} ty >= 20 ty)")
    elif any(x in note_lower for x in ["tài chính mạnh", "tai chinh manh", "không thành vấn đề", "khong thanh van de", "tài chính cao", "tai chinh cao"]):
        score += 50
        extracted["budget"] = "Tai chinh manh"
        reasons_plus.append("Ngan sach lon (Co de cap 'tai chinh manh' / 'khong thanh van de')")

    # Loại hình cao cấp
    luxury_types = ["biệt thự đơn lập", "biet thu don lap", "biệt thự", "biet thu", "penthouse", "shophouse mặt đường lớn", "shophouse mat duong lon", "shophouse", "quỹ đất công nghiệp", "quy dat cong nghiep", "sàn văn phòng diện tích lớn", "san van phong dien tich lon", "sàn văn phòng", "san van phong"]
    for t in luxury_types:
        if t in note_lower:
            score += 50
            extracted["property_type"] = t.title()
            reasons_plus.append(f"Loai hinh cao cap ({t.title()})")
            break

    # Vị trí đắc địa
    prime_locations = ["quận 1", "quan 1", "ven sông", "ven song", "vinhomes ocean park", "ocean park", "phú mỹ hưng", "phu my hung"]
    for l in prime_locations:
        if l in note_lower:
            score += 50
            extracted["location"] = l.title()
            reasons_plus.append(f"Vi tri dac dia ({l.title()})")
            break

    # Đối tượng khách hàng
    vip_clients = ["chủ doanh nghiệp", "chu doanh nghiep", "nhà đầu tư chuyên nghiệp", "nha dau tu chuyen nghiep", "mua sỉ", "mua si", "mua số lượng lớn", "mua so luong lon"]
    for c in vip_clients:
        if c in note_lower:
            score += 50
            reasons_plus.append(f"Doi tuong VIP ({c.capitalize()})")
            break

    # Tính cấp thiết & Minh bạch
    urgency_keywords = ["pháp lý chuẩn 100%", "phap ly chuan 100%", "sổ hồng riêng", "so hong rieng", "gặp trực tiếp chủ đầu tư để đàm phán", "gap truc tiep chu dau tu de dam phan", "gặp trực tiếp chủ đầu tư", "gap truc tiep chu dau tu"]
    for u in urgency_keywords:
        if u in note_lower:
            score += 50
            extracted["urgency"] = "Cao"
            reasons_plus.append("Cap thiet & Minh bach (Yeu cau phap ly/Muon gap truc tiep CDT)")
            break


    # --- 2. TIÊU CHÍ TRỪ 50 ĐIỂM (RÁC) ---
    
    # Yêu cầu phi thực tế (Ví dụ: Nhà Quận 1 giá 1-2 tỷ, nhà trung tâm bể bơi giá vài trăm triệu)
    if ("quận 1" in note_lower or "quan 1" in note_lower) and budget_match:
        budget_val = int(budget_match.group(1))
        if budget_val <= 3:
            score -= 50
            reasons_minus.append(f"Yeu cau phi thuc te (Mua nha Q1 voi gia {budget_val} ty)")
    elif ("trung tâm" in note_lower or "trung tam" in note_lower or "sân vườn" in note_lower or "san vuon" in note_lower or "hồ bơi" in note_lower or "ho boi" in note_lower) and budget_match:
        budget_val = int(budget_match.group(1))
        if budget_val < 1:  # Dưới 1 tỷ
            score -= 50
            reasons_minus.append("Yeu cau phi thuc te (Nha trung tam co san vuon ho boi gia qua thap)")
    elif "vài trăm triệu" in note_lower or "vai tram trieu" in note_lower:
        score -= 50
        reasons_minus.append("Yeu cau phi thuc te (Gia qua thap so voi thi truong)")

    # Không có nhu cầu
    if any(x in note_lower for x in ["nhầm số", "nham so", "không có nhu cầu", "khong co nhu cau", "dữ liệu cũ", "du lieu cu", "nhầm ngành", "nham nganh"]):
        score -= 50
        reasons_minus.append("Khong co nhu cau (Bao nham so/nganh/du lieu cu)")

    # Khách hàng không thiện chí
    if any(x in note_lower for x in ["hỏi giá cho vui", "hoi gia cho vui", "chưa có ý định mua", "chua co y dinh mua", "thái độ không hợp tác", "thai do khong hop tac"]):
        score -= 50
        reasons_minus.append("Khach hang thieu thien chi (Hoi cho vui/Khong hop tac)")

    # Spam/Quảng cáo
    if any(x in note_lower for x in ["bảo hiểm", "bao hiem", "vay vốn", "vay von", "mời chào dịch vụ", "moi chao dich vu", "tuyển dụng", "tuyen dung"]):
        score -= 50
        reasons_minus.append("Spam/Quang cao (Chao moi bao hiem, vay von, dich vu khac)")

    # Thông tin liên lạc lỗi
    if any(x in note_lower for x in ["thuê bao", "thue bao", "gọi nhiều lần không bắt máy", "goi nhieu lan khong bat may", "không nghe máy", "khong nghe may", "không phản hồi zalo", "khong phan hoi zalo"]):
        score -= 50
        reasons_minus.append("Loi lien lac (Thue bao/Khong bat may/Khong rep Zalo)")


    # --- 3. ĐIỀU CHỈNH ĐIỂM SỐ & PHÂN LOẠI ---
    
    # Boundary correction
    if len(reasons_plus) > 0 and len(reasons_minus) == 0:
        score = min(score, 100)
    elif len(reasons_minus) > 0:
        score = max(score - 50, 0)
    else:
        score = 50  # Neutral baseline

    # Classification
    if score >= 90:
        classification = "VIP"
    elif score <= 20:
        classification = "Rác"
    else:
        classification = "Tiềm năng"

    # Assemble Vietnamese reason string
    reason_parts = []
    if reasons_plus:
        reason_parts.append("[CONG DIEM: VIP] " + ", ".join(reasons_plus))
    if reasons_minus:
        reason_parts.append("[TRU DIEM: RAC] " + ", ".join(reasons_minus))
    
    if not reason_parts:
        # Check if they are normal group
        if "chung cư" in note_lower or "chung cu" in note_lower or "nhà phố" in note_lower or "nha pho" in note_lower:
            reason_str = "Khach hang tim mua chung cu/nha pho tam trung (3-10 ty) - Giu nguyen diem co so."
        elif "vay" in note_lower or "ngân hàng" in note_lower or "ngan hang" in note_lower:
            reason_str = "Khach hang can vay ngan hang, dang can nhac chinh sach - Giu nguyen diem co so."
        else:
            reason_str = "Nhu cau thuc te, can tu van them ve phap ly/vi tri - Giu nguyen diem co so."
    else:
        reason_str = " | ".join(reason_parts)

    return {
        "score": score,
        "classification": classification,
        "reason": reason_str,
        "budget": extracted["budget"],
        "property_type": extracted["property_type"],
        "location": extracted["location"],
        "urgency": extracted["urgency"]
    }


def write_to_excel(leads_data):
    """
    Exports the scored leads into a premium, styled Excel file.
    """
    print("\n==================================================")
    print("3. KET XUAT FILE EXCEL BAO CAO")
    print("==================================================")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "AI Lead Scoring Results"
    ws.views.sheetView[0].showGridLines = True
    
    # Font & Styles
    font_name = "Segoe UI"
    title_font = Font(name=font_name, size=15, bold=True, color="1E3A8A")
    subtitle_font = Font(name=font_name, size=10, italic=True, color="555555")
    header_font = Font(name=font_name, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Dark Navy
    
    cell_font = Font(name=font_name, size=10)
    bold_cell_font = Font(name=font_name, size=10, bold=True)
    
    fill_vip = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")     # Emerald Green Accent
    font_vip = Font(name=font_name, size=10, bold=True, color="065F46")
    
    fill_normal = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # Amber Gold Accent
    font_normal = Font(name=font_name, size=10, bold=True, color="92400E")
    
    fill_trash = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")   # Coral Red Accent
    font_trash = Font(name=font_name, size=10, bold=True, color="991B1B")

    center_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=False)
    
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    # Title Rows
    ws.merge_cells("A1:K1")
    ws["A1"] = "DANH SACH KHACH HANG DA CHAM DIEM - AI LEAD SCORING"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    
    ws.merge_cells("A2:K2")
    ws["A2"] = "He thong tu dong phan tich & xep hang khach hang dua tren tieu chi nghiep vu"
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    
    ws.append([]) # Empty spacing row
    
    # Headers
    headers = [
        "Mã KH", "Họ tên", "Số điện thoại", "Nhu cầu chi tiết", 
        "Điểm số", "Phân loại", "Lý do chi tiết", 
        "Ngân sách", "Loại hình BĐS", "Vị trí yêu cầu", "Độ cấp thiết"
    ]
    ws.append(headers)
    header_row = 4
    
    # Format Headers
    for c_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=c_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    # Append Data
    for lead in leads_data:
        ws.append([
            lead["id"],
            lead["name"],
            lead["phone"],
            lead["note"],
            lead["score"],
            lead["classification"],
            lead["reason"],
            lead["budget"],
            lead["property_type"],
            lead["location"],
            lead["urgency"]
        ])
        
    # Format Data Rows
    start_row = 5
    end_row = ws.max_row
    
    for r in range(start_row, end_row + 1):
        ws.row_dimensions[r].height = 20
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = cell_font
            cell.border = thin_border
            cell.alignment = left_align
            
        # Alignment Overrides
        ws.cell(row=r, column=1).alignment = center_align # ID
        ws.cell(row=r, column=3).alignment = center_align # SĐT
        ws.cell(row=r, column=5).alignment = center_align # Score
        ws.cell(row=r, column=6).alignment = center_align # Classification
        ws.cell(row=r, column=11).alignment = center_align # Urgency
        
        # Classification Conditional Colors
        score_cell = ws.cell(row=r, column=5)
        score_cell.font = bold_cell_font
        
        class_cell = ws.cell(row=r, column=6)
        val = class_cell.value
        if val == "VIP":
            class_cell.fill = fill_vip
            class_cell.font = font_vip
        elif val == "Tiềm năng":
            class_cell.fill = fill_normal
            class_cell.font = font_normal
        elif val == "Rác":
            class_cell.fill = fill_trash
            class_cell.font = font_trash
            
    # Auto-adjust Column Widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        for idx, cell in enumerate(col):
            if idx < 3: # Skip title blocks
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
                
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Override wide columns manually for readability
    ws.column_dimensions['D'].width = 40  # Nhu cầu
    ws.column_dimensions['G'].width = 50  # Lý do
    
    ws.row_dimensions[4].height = 26
    
    # Save file
    try:
        wb.save(OUTPUT_FILE)
        print(f"-> Luu file excel thanh cong: '{OUTPUT_FILE}'")
        print("-> File da duoc dinh dang dep mat, san sang de ban giao.")
    except Exception as e:
        print(f"-> Loi khi ghi file Excel: {e}")


def main():
    print("==================================================")
    print(" AI LEAD SCORING SYSTEM - AUTOMATION CLI ")
    print("==================================================")
    
    # 1. Fetch
    df = fetch_data()
    
    # 2. Check and map columns
    col_name = None
    col_phone = None
    col_note = None
    
    for col in df.columns:
        col_lower = str(col).lower()
        if ('tên' in col_lower or 'ho ten' in col_lower or 'name' in col_lower) and not ('nhu cầu' in col_lower or 'nhu cau' in col_lower or 'note' in col_lower or 'ghi chú' in col_lower):
            col_name = col
        if 'thoại' in col_lower or 'sđt' in col_lower or 'sdt' in col_lower or 'phone' in col_lower or 'liên hệ' in col_lower:
            col_phone = col
        if 'nhu cầu' in col_lower or 'nhu cau' in col_lower or 'ghi chú' in col_lower or 'ghi chu' in col_lower or 'note' in col_lower or 'inquiry' in col_lower or 'mô tả' in col_lower or 'mo ta' in col_lower:
            col_note = col

    # Fallback to defaults
    if not col_name: col_name = df.columns[1] if len(df.columns) > 1 else df.columns[0]
    if not col_phone: col_phone = df.columns[2] if len(df.columns) > 2 else df.columns[0]
    if not col_note: col_note = df.columns[3] if len(df.columns) > 3 else df.columns[0]
    
    print(f"-> Nhan dien cac cot du lieu: Ho ten = '{col_name}', SDT = '{col_phone}', Nhu cau = '{col_note}'")
    
    # 3. Score leads
    print("\n==================================================")
    print("2. CHAY AI LEAD SCORING (RULE-BASED ENGINE)")
    print("==================================================")
    print("Dang phan tich va tu dong cham diem tung khach hang...")
    
    leads_results = []
    for idx, row in df.iterrows():
        name = str(row[col_name]) if pd.notna(row[col_name]) else "Khong ten"
        phone = str(row[col_phone]) if pd.notna(row[col_phone]) else "Khong co"
        note = str(row[col_note]) if pd.notna(row[col_note]) else ""
        
        # Core heuristic analysis
        res = heuristic_scoring(note)
        
        leads_results.append({
            "id": idx + 1,
            "name": name,
            "phone": phone,
            "note": note,
            "score": res["score"],
            "classification": res["classification"],
            "reason": res["reason"],
            "budget": res["budget"],
            "property_type": res["property_type"],
            "location": res["location"],
            "urgency": res["urgency"]
        })
        
        print(f"Lead {idx+1}: {name} - Diem: {res['score']} -> {res['classification']}")
        
    # Summary of stats
    total = len(leads_results)
    vips = sum(1 for l in leads_results if l["classification"] == "VIP")
    potentials = sum(1 for l in leads_results if l["classification"] == "Tiềm năng")
    trash = sum(1 for l in leads_results if l["classification"] == "Rác")
    print(f"\n-> Hoan thanh: Tong so {total} khach hang (VIP: {vips}, Tiem nang: {potentials}, Rac: {trash})")
    
    # 4. Write Excel
    write_to_excel(leads_results)
    
    print("\n==================================================")
    print(" CHUONG TRINH HOAN THANH TOT DEP ")
    print("==================================================")


if __name__ == "__main__":
    main()
